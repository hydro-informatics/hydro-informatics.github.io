#!/usr/bin/env python
# coding: utf-8

# # Manipulation of Workbooks and JSON Files
# 
# Summary: Create, manipulate and copy xlsx-workbooks and JSON files.
# 
# This lesson starts with some background information about XML and what XML has to do with workbooks and JSON: XML is the abbreviation of [*E**x**tensible **M**arkup **L**anguage*](https://www.w3.org/TR/xml/) that defines rules for encoding documents. XML was designed for straightforward usage over the internet and we encounter XML documents all the time, on websites (e.g., *XHTML*), in the shape office documents (e.g., Office Open XML such as *docx*, *pptx*, or *xlsx*), or podcasts (e.g., *RSS*). The strength of the XML format is its characteristic of being both machine-readable (i.e., a computer can process XML files) and human-readable (i.e., we can read it like a newspaper). In simple words, entering formulas in an xlsx workbook is simultaneously machine-readable and human-readable, since humans and computers can interpret and evaluate the formulas in this XML frame. Other file formats such as [*JSON* (JavaScript Object Notation)](https://www.json.org/json-en.html) resemble XML and this is why we look at how *Python* can extract information from and export information to both formats. In water resources engineering and research, we are mainly interested in the exchange of information with office workbooks (*xlsx* files), or with JSON files, which provide boundary conditions for numerical models.
# 
# ## Workbook (xlsx) Handling
# 
# Why do we want to communicate with workbooks at all? We have already seen that *Python* is much more powerful than office programs for the systematic analysis of data sets. However, *Python* requires the abstraction of data fields in our minds to visualize for example the structure of a nested list. For this reason, data from and for marketing, your boss, or public authorities are often required to have visually easy-to-use workbook formats, which can be overlooked quickly. Still we want to leverage the content of such workbook information efficiently with *Python* and we want to produce visually simplistic output that anyone can read without any *Python* knowledge.
# 
# We have already seen that *pandas* provides easy routines for importing and exporting data from and to workbooks, respectively ([cf. the data processing & file handling section](pynum.html#pd-files)). *pandas* uses [*XlsxWriter*](https://xlsxwriter.readthedocs.io/) or [*openpyxl*](https://openpyxl.readthedocs.io/en/stable/) depending on what is available in the current *Python* environment*. Both packages have different behaviors, in particular when it comes to writing new or in existing workbooks. This is why the most robust approach is the direct use of one particular workbook handling library, where *openpyxl* is one of the most powerful options (note: this assertion is subjective) and this section introduces workbook handling with *openpyxl*.
# 
# This introduction uses the following workbook-related terms:
# 
# * **workbook** is the main *xlsx* file we work with (also called *spreadsheet*);
# * **sheet** is tabular content of a workbook and one workbook can have multiple sheets;
# * **column**s are vertical lines in a sheet;
# * **row**s are horizontal lines in a sheet;
# * **cell**s are elements of a sheet.
# 
# ### Create a Workbook
# *openpyxl* has a `Workbook` class that enables to create and fill workbooks with data. Typically, an instance of the `Workbook` class is called `wb` and worksheets contain the letters `ws`.

# In[1]:


import openpyxl as oxl
wb = oxl.Workbook()  # create a Workbook instance
ws = wb.active  # activate worksheet
ws.titel = "Gaussian 2D"  # name worksheet
ws["A1"] = "Gaussian sample data"  # write to cell A1

# generate some data
x, y = np.meshgrid(np.linspace(-1, 1, 20), np.linspace(-1, 1, 20))
dis = np.sqrt(x * x + y * y)
sigma, mu = 1.0, 0.0
gaussian = np.exp(-((dis - mu) ** 2 / (2.0 * sigma ** 2)))

# write data to worksheet
m, n = gaussian.shape
for i in range(1, m):
    for j in range(1, n):
        _ = ws.cell(row=i+1, column=j, value=gaussian[i-1, j-1])

print("Workbook data in cell A2: "+ str(ws["A2"].value))
print("Corresponds to np.array value: " + str(gaussian[0, 0]))

# save and close (destruct object) workbook
wb.save(filename="data/python_workbook.xlsx")
wb.close()


# ![img](https://raw.githubusercontent.com/sschwindt/hydroinformatics/main/docs/img/py-xlsx.png)

# ### Read and Manipulate an Existing Workbook
# 
# ```{warning}
# When *openpyxl* opens an existing workbook, it cannot read graphical objects (e.g., graphs, shapes, or images). For this reason, saving the workbook with the same name will make graphical objects disappear.
# ```
# 
# *openpyxl* reads existing workbooks with `openpyxl.load_workbook(filename=str())`. This function accepts optional keyword arguments, where the most important are:
# 
# * `read_only=BOOLEAN` decides weather a workbook is opened in *read only* mode. A workbook can only be manipulated if `read_only=False` (this is the default option, which can be useful to handle large files or to ensure that graphical objects are not lost).
# * `write_only=BOOLEAN` decides weather a workbook is opened in *write only* mode. If `write_only=True` (the default is `False`), no data can be read from a workbook, but writing data is significantly faster (i.e., this option is useful to write large datasets).
# * `data_only=BOOLEAN` determines if cell formulae or cell data will be read. For example, when a workbook cell's content is `=PI()`, `data_only=False` (this is the default option) reads the cell value as `=PI()` and `data_only=True` reads the cell value as `3.14159265359`.
# * `keep_vba=BOOLEAN` controls whether *Visual Basic* elements (macros) are kept or not. The default is `keep_vba=False` (i.e., no preservation) and `keep_vba=True` will still not enable a modification of Visual Basic elements.
# 
# If `read_only=False`, we can manipulate cell values and also cell formats, including data formats (e.g., date, time, and [many more](https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html)), [font properties (and many more cell styles)](https://openpyxl.readthedocs.io/en/stable/styles.html), or colors in *HEX Color Code* ([find your favorite color here](https://www.colorcodehex.com/)). The following example opens the above created `python_workbook.xlsx`, adds a new worksheet, illustrates the implementation of some cell styles and fills it with randomized discharge measurements.

# In[2]:


import datetime
from openpyxl.styles import Font, Alignment, PatternFill
wb = oxl.load_workbook(filename="data/python_workbook.xlsx", read_only=False)
ws = wb.create_sheet(title="Discharge")

# define title styles
title_font = Font(name="Tahoma", size="11", bold=True, italic=True, color="C1D0DE")
title_fill = PatternFill(fill_type="solid", start_color="050505", end_color="073AD4")
title_align = Alignment(horizontal='center', vertical='bottom', text_rotation=0,
                        wrap_text=False, shrink_to_fit=False, indent=0)

date_time_format = "yyyy-mm-dd hh:mm:ss"
ws["A1"] = "Date-Time (%s)" % date_time_format

title_cell_flow = ws["B1"]
title_cell_flow.value = "Discharge (CMS)"
title_cell_flow.font = title_font
title_cell_flow.fill = title_fill
title_cell_flow.alignment = title_align

# define time period and time delta of 1 hour = 3600 seconds
current_date_time = datetime.datetime(2040, 12, 24, 0, 0)
dt = datetime.timedelta(seconds=3600)

# write random discharges to workbooks
for row in ws.iter_rows(min_row=2, max_row=26, min_col=1, max_col=2):
    row[0].value = current_date_time
    row[0].number_format = date_time_format
    row[1].value = np.random.random_sample(size=None) * 100
    row[1].number_format = "0.00"
    current_date_time += dt
    
wb.save("data/python_workbook_reloaded.xlsx")
wb.close()


# ![img](https://raw.githubusercontent.com/sschwindt/hydroinformatics/main/docs/img/py-xlsx-reloaded.png)

# The below code block provides the short helper function `read_columns` to read only one or more columns into a (nested) *list* (reads until the maximum number of rows, defined by `ws.rows`, in a workbook is reached). A similar function can be written for reading rows.

# In[3]:


def read_columns(ws, start_row=0, columns="ABC"):
    return [ws["{}{}".format(column, row)].value for row in range(start_row, int(ws.rows.__sizeof__()) + 1) for column in columns]

# example usage:
wb = oxl.load_workbook(filename="data/python_workbook.xlsx", read_only=False)
ws = wb.active
col_D = read_columns(ws, start_row=2, columns="D")
col_F = read_columns(ws, start_row=2, columns="F")
wb.close()


# ### Formulae in Workbooks
# 
# The optional keyword argument `data_only=False` enables reading workbook formula instead of cell values. However, not all workbook formulae are recognized by *openpyxl* and in the case of doubts, a dirty try-and-error approach is the only remedy. As an example, change `SQRT` in the below example to the formula in question.

# In[4]:


from openpyxl.utils import FORMULAE
print("SQRT" in FORMULAE)


# ### (Un)merge Cells
# 
# Merging and un-merging cells is a popular office function for style purposes and *openpyxl* also provides function to perform merge operations:

# In[5]:


ws.merge_cells(start_row=1, end_row=3, start_column=1, end_column=2)
ws.unmerge_cells(start_row=1, end_row=3, start_column=1, end_column=2)


# ### Charts (Plots)
# 
# In the unlikely event that you want to insert plots directly into workbooks ([`matplotlib`](pyplot.html#matplotlib) is way more powerful), *openpyxl* offers features for this purpose as well. To illustrate the creation of a an area chart, the below code block re-uses the first column of random values in the  previously created `python_workbook.xlsx`. 

# In[6]:


from openpyxl.chart import AreaChart, Reference, Series

wb = oxl.load_workbook(filename="data/python_workbook.xlsx", read_only=False)
ws = wb.active

chart = AreaChart()
chart.title = "Random Gaussian"
chart.style = 10
chart.x_axis.title = "Cell row"
chart.y_axis.title = "Random value (-)"

col_D = Reference(ws, min_col=4, min_row=2, max_row=20)
col_F = Reference(ws, min_col=6, min_row=2, max_row=20)

chart.add_data(col_F, titles_from_data=False)
chart.add_data(col_D, titles_from_data=False)

ws.add_chart(chart, "B2")

wb.save("data/python_workbook_chart.xlsx")
wb.close()


# ![img](https://raw.githubusercontent.com/sschwindt/hydroinformatics/main/docs/img/py-xlsx-plot.png)

# Other workbook charts are available and their implementation (still: why would you?) is explained in the [*openpyxl* docs](https://openpyxl.readthedocs.io/en/stable/charts/introduction.html).

# ### Customize Workbook Manipulation
# There are many ways of modifying workbooks and *openpyxl* provides close-to "shovel-ready" methods to manipulate workbooks. Still, in order to avoid re-reading this lesson every time you want to manipulate a workbook, it is much more convenient to have your own workbook manipulation classes ready. For example, use custom `Read` and `Write` classes, where `Read` is the parent class of the `Write` class (see the section on [inheritance of classes](classes.html#inheritance)). The `Read` class may contain tailored functions for reading specific columns, rows, or arrays. The below code block illustrates a basic example for such `Read` and `Write` classes with the above `read_columns` function implemented as a method of the `Read` class.

# In[7]:


import openpyxl as oxl

class Read:
    def __init__(workbook_name="", *args, **kwargs):
        read_only = kwargs.get("read_only")
        data_only = kwargs.get("data_only")
        sheet_name = kwargs.get("data_only")
        self.wb = oxl.load_workbook(filename=workbook_name, read_only=read_only, data_only=data_only)
        if sheet_name:
            self.ws = self.wb.worksheets[worksheet]
        else:
            self.ws = self.wb[self.wb.sheetnames[0]]
            
    def read_columns(self, start_row=0, columns="ABC"):
        return [self.ws["{}{}".format(column, row)].value for row in range(start_row, len(self.ws.rows) + 1) for column
                in columns]
            
    def __call__(self):
        print(dir(self))

            
class Write(Read):
    def __init__(workbook_name="", *args, **kwargs):
        data_only = kwargs.get("data_only")
        sheet_name = kwargs.get("data_only")
        Read.__init__(workbook_name=workbook_name, read_only=False, data_only=data_only, sheet_name=sheet_name)


# An extended example script with more complex `Read` and `Write` classes can be downloaded from the [course repository](https://github.com/hydro-informatics/material-py-codes/raw/master/workbooks/xlsx.py) (please note that this script is only temporary available when the course takes place).
# 
# ```{admonition} Challenge
# What are your favorite fonts, table colors and layouts? Write your own `Read` and `Write` classes with formatting methods to have a personal template ready to be used at any time.
# ```

# ### An Example from Water Resources Engineering and Research
# 
# The ecological restoration or enhancement of rivers requires, among other data, information on preferred water depths and flow velocities of target fish species. This information is established by biologists and then often provided in the shape of so-called [habitat suitability index (HSI)](https://riverarchitect.github.io/RA_wiki/SHArC#hefish) curves in workbook formats. As water resources researchers and engineers, we produce geospatially explicit data of water depth and flow velocity with numerical models. The output of two- or three-dimensional numerical models is way too large for being handled with office applications. So we need an advanced tool such as *Python* to handle the geospatially explicit data and to read and interpolate HSI curves from workbooks. How does that look like technically? The course assignments will help you to find out ...

# ```{admonition} Exercise
# Get more familiar with workbook handling in the [Sediment transport (1d)](../exercises/ex-sediment) exercise.
# ```

# ## JSON
# 
# JavaScript Object Notation ([JSON](https://www.json.org/json-en.html)) files have a similar structure to XML and enable the structured storage of (human-readable) data. For example, the numerical code *BASEMENT v.3.x* ([read more in the numerical modelling chapter](../numerics/basement)) uses a *model.json* and a *simulation.json* file to store model setup parameters such as material properties. In water resources engineering and research, we often want to automate running numerical models, which involves the optimization of model parameters stored in *json* files. This is where *Python* steps in with the `JSON` package and `pandas`' *JSON* modules.

# ### *JSON* file structure
# 
# A *JSON* file consists of two types of data structures, which are *dictionary* objects and *array*s in the form of *lists* of values. The *dictionary* objects in a *JSON* file correspond to the same format that we already know in *Python*: Pairs of keys (names) and values embraced by curly brackets (*braces*) `{"name": value}`. The `value` can be a *string*, *numeric*, a comm-separated *list* `[]` (*array*) of data, or another *dictionary*.
# The following example shows a *JSON* file called `river_struct.json` with a `RIVER` key that has a nested dictionary as value. The value-*dictionary* contains three keys (`NAME`, `GEOMETRY`, and `HYDRAULICS`.
# 
# > ***Tip***: Take a couple of minutes to understand the elements of `river_struct.json`.<br>
# What is the purpose of the `FLOWBOUNDARIES` in `GEOMETRY`? <br>
# How could the `FLOWBOUNDARIES` be related to the `BOUNDARY` key of `HYDRAULICS`?<br>
# What units could the `FRICTION` values correspond to?<br>
# Can you find the river on a map?

# In[8]:


{
	"RIVER": {
		"NAME": "Vanilla Flow",
		"GEOMETRY": {
			"REGIONS": [
				{
				  "type": "wet",
				  "name": "riverbed"
				},
				{
				  "type": "dry",
				  "name": "floodplain"
				}
			],
			"FLOWBOUNDARIES": [
				{
				  "name": "Inflow",
				  "nodes": [1, 3, 7, 31]
				},
				{
				  "name": "Outflow",
				  "nodes": [89, 90, 76, 69, 95]
				}
			]
		},
		"HYDRAULICS": {
			"BOUNDARY": [
				{
					"discharge_file": "/simulation/directory/Inflow.txt",
					"name": "Inflow",
					"slope": 0.005,
					"type": "hydrograph"
				},
				{
					"name": "Outflow",
					"type": "zero_gradient"
				}
			],
			"FRICTION": {
				"cobble": 20.0,
				"gravel": 26.0,
				"sand": 41
			}
		},
		"LOCATION": [48.744079, 9.103928]
	}
}


# ### Read (Decode) and Write (Encode) *JSON* Files with the `json` Library 
# 
# *JSON* files can be implemented in many programming languages including *HTML* and *Python*. This is also the reason why *Jupyter* notebooks (as used in this course) can be run in *Python* and displayed as a web page. *Python* has a built-in `json` library that enables *JSON* decoding and encoding. The `json` library provides a `json.dumps(DATA)` method to "dump" (i.e., encode) data in *JSON* format. Vice versa, the `json.load()` method reads data from *JSON* files. The following example illustrates encoding and decoding an arbitrary nested dataset with the `json` library.

# In[9]:


import json
# create arbitrary nested data (list, dictionary, tuple)
data_for_json = ["list_element1", {"dict_key": ("tuple_element", "text", 1.0, None)}]

# create a json file
json_file = open("data/my-first.json", mode="w+")
# encode the random nested data list in json format and write to file
json_file.write(json.dumps(data_for_json))
# close file
json_file.close()

# re-open the json file to read data
with open("data/my-first.json", mode="r") as re_opened_file:
    raw_data = re_opened_file.readline()

# decode json data in a Python variable
data_from_json = json.loads(raw_data)
print(json.dumps(data_from_json))


# The [*Python* docs](https://docs.python.org/3/library/json.html) provide more options and descriptions on using the `json` library. However, here we will (once again) make use of the *pandas* library, which offers some powerful features for handling json data.

# (read-json)=
# ### Read (Decode) and Write (Encode) *JSON* Files with *pandas* 
# 
# *pandas* (recall [data and file handling](pynum.html#pandas)) enables reading *JSON* files into its convenient table format with an embedded usage of the `json` library.
# The following code block uses the [`pandas.read_json(FILE)`](https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_json.html) method to read the above shown `RIVER` sample file.

# In[10]:


river = pd.read_json("data/river_struct.json")
print(river)


# Since a river without data is like ice cream without taste, we will add (random) data of flow characteristics to the data structure. Let us assume that we have used the data from `river_struct.json` to simulate a stationary discharge in a two-dimensional numerical model. As a result we have two regular grids (arrays) with data on flow velocity and flow depth. Now, we want to append both the flow depth and flow velocity arrays in the form of a result structure (dictionary) in the `river_struct.json` and give the river a new name.

# In[11]:


# create random data
h = np.random.weibull(np.arange(0,100)).reshape(10, 10)
u = np.random.weibull(np.arange(0,100)).reshape(10, 10)

# append RESULTS row to pandas dataframe
river_dict = river.to_dict()
river_dict["RIVER"].update({"RESULTS": {"flow_depth": h, "flow_velocity": u}})
updated_river = pd.DataFrame.from_dict(river_dict)

# re-NAME RIVER
updated_river["RIVER"]["NAME"] = "Honey river"
print(updated_river)

# export to JSON
updated_river.to_json("data/river_results.json")


# ![img](https://raw.githubusercontent.com/sschwindt/hydroinformatics/main/docs/img/py-json-file.png)

# ```{admonition} Exercise
# Get more familiar with *JSON* file handling in the [geospatial ecohydraulics](../exercises/ex-geco) exercise (requires understanding the chapter on geospatial *Python*).
# ```
